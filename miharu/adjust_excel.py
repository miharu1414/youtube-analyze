'''
    sheet_width.py
    purpose: make new xlsx and set width automatically
'''
import openpyxl as xl
from openpyxl.styles import Alignment  

def Adjust_width(filepath):
    # set input file name
    inputfile = filepath

    # read input xlsx
    wb1 = xl.load_workbook(filename=inputfile)
    ws1 = wb1.worksheets[0]
    
    #最大行
    maxRow = ws1.max_row + 1

    #最大列
    maxClm = ws1.max_column + 1
    
    col = list(ws1.columns)


    #列ループ
    for j in range(1,maxClm):
        max_length = 0
        #行を逆ループ
        for i in reversed(range(1,maxRow)):
            if len(str(ws1.cell(i,j).value)) > max_length:
                max_length = len(str(ws1.cell(i,j).value))
            ws1.cell(i,j).alignment = Alignment(vertical='justify')
                 
        ws1.column_dimensions[col[j-1][0].column_letter].width = min(30,max_length)
            
            
            

    # save xlsx file
    wb1.save(inputfile)

if __name__ == '__main__':
    filepath = "data\gnIOzY7esA0.xlsx"
    Adjust_width(filepath)