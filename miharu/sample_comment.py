from xmlrpc.client import NOT_WELLFORMED_ERROR

from miharu.adjust_excel import Adjust_width

import requests
import json

import os
from dotenv import load_dotenv
import openpyxl

load_dotenv('.env') 

URL = 'https://www.googleapis.com/youtube/v3/'
# ここにAPI KEYを入力
API_KEY = os.environ.get("API_KEY")
# ここにVideo idを入力
VIDEO_ID = 'N43buH_8060'
# VIDEO_URL = 'N43buH_8060'


def Make_comment_table(channel_id,video_id):
    """URLを入力として受け取り、その動画に投稿されたコメント一覧をcsvに出力します。"""
    # xlsxに保存
    # video_id = Url_to_id(video_url)
 
    file_path = f'data/{channel_id}/comment/{video_id}.xlsx'
  
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]
    no = 1
    
    def print_video_comment(no, video_id, next_page_token):
        if no == 1:
            columns = ['id','text','like_cnt','reply_cnt','user_name','parentId']
            for j in range(len(columns)):
                ws.cell(1,j+1).value = columns[j]
        i = no + 1
        
        params = {
            'key': API_KEY,
            'part': 'snippet',
            'videoId': video_id,
            'order': 'relevance',
            'textFormat': 'plaintext',
            'maxResults': 100,
        }

        if next_page_token is not None:
            params['pageToken'] = next_page_token
        response = requests.get(URL + 'commentThreads', params=params)
        resource = response.json()
    

        
        
        for comment_info in resource['items']:

            # コメント
            text = comment_info['snippet']['topLevelComment']['snippet']['textDisplay']
            # グッド数
            like_cnt = comment_info['snippet']['topLevelComment']['snippet']['likeCount']
            # 返信数
            reply_cnt = comment_info['snippet']['totalReplyCount']
            # ユーザー名
            user_name = comment_info['snippet']['topLevelComment']['snippet']['authorDisplayName']
            # Id
            parentId = comment_info['snippet']['topLevelComment']['id']
            print('{:0=4}\t{}\t{}\t{}\t{}'.format(no, text.replace('\r', '\n').replace('\n', ' '), like_cnt, user_name, reply_cnt))
    
            ws.cell(i,1).value = no
            ws.cell(i,2).value = text
            ws.cell(i,3).value = like_cnt
            ws.cell(i,4).value = reply_cnt
            ws.cell(i,5).value = user_name
            ws.cell(i,6).value = parentId
            if reply_cnt > 0:
                cno = 1
                print_video_reply(no, cno, video_id, None, parentId)
            no = no + 1
            i += 1

        if 'nextPageToken' in            resource:
            print_video_comment(no, video_id, resource["nextPageToken"])




    def print_video_reply(no, cno, video_id, next_page_token, id):
        params = {
            'key': API_KEY,
            'part': 'snippet',
            'videoId': video_id,
            'textFormat': 'plaintext',
            'maxResults': 50,
            'parentId': id,
        }

        if next_page_token is not None:
            params['pageToken'] = next_page_token
        response = requests.get(URL + 'comments', params=params)
        resource = response.json()

        for comment_info in resource['items']:
            # コメント
            text = comment_info['snippet']['textDisplay']
            # グッド数
            like_cnt = comment_info['snippet']['likeCount']
            # ユーザー名
            user_name = comment_info['snippet']['authorDisplayName']

            print('{:0=4}-{:0=3}\t{}\t{}\t{}'.format(no, cno, text.replace('\r', '\n').replace('\n', ' '), like_cnt, user_name))
            cno = cno + 1

        if 'nextPageToken' in resource:
            print_video_reply(no, cno, video_id, resource["nextPageToken"], id)


    print_video_comment(no, video_id, None)

    wb.save(file_path)
    
    Adjust_width(file_path)
    
    
    
 
 
def Url_to_id(url):
    target = 'watch?v='
    idx = url.find(target)
    id = url[idx+1:]
    return id
    
 
if __name__ == '__main__':
    # コメントを全取得する
    video_id = VIDEO_ID

    Make_comment_table(video_id)

