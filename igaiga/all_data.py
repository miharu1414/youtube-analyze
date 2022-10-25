import urllib.request
import urllib.parse
import json
import csv
import isodate
import datetime
import openpyxl

APIKEY = API_KEY
#↓分析したいチャンネルのidを入力
channel_id = 'UC4B6r1TQyN5LhtDk-aaA9Qg'

dt_now = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
nextPageToken = ''
item_count = 0
outputs = []
a = ['publishedAt', 'title', 'description', 'id', 'thumbnail_url', 'categoryId', 'liveBroadcastContent', 'duration', 'viewCount', 'likeCount', 'favoriteCount', 'commentCount', 'embedHtml']
outputs.append(['publishedAt', 'title', 'description', 'url', 'thumbnail_url', 'categoryId', 'liveBroadcastContent', 'duration', 'viewCount', 'likeCount', 'favoriteCount', 'commentCount', 'embedHtml'])
n = 0
wb = openpyxl.Workbook()
ws = wb.worksheets[0]

for j in range(len(a)):
    ws.cell(1, j+1).value = a[j]

while True:
    #searchメソッドでvideoid一覧取得
    param = {
        'part':'snippet',
        'channelId':channel_id,
        'maxResults':50,
        'order':'date',
        'type':'video',
        'pageToken':nextPageToken,
        'key':APIKEY
    }
    target_url = 'https://www.googleapis.com/youtube/v3/search?'+urllib.parse.urlencode(param)
    print('動画リスト取得')
    print(target_url)
    req = urllib.request.Request(target_url)
    try:
        with urllib.request.urlopen(req) as res:
            search_body = json.load(res)
            item_count += len(search_body['items'])
            video_list = []
            for item in search_body['items']:
                #videoメソッド用list作成
                video_list.append(item['id']['videoId'])
                
            #videoメソッドで動画情報取得
            param = {
                'part':'id,snippet,contentDetails,liveStreamingDetails,player,recordingDetails,statistics,status,topicDetails',
                'id':",".join(video_list),
                'key':APIKEY
            }
            target_url = 'https://www.googleapis.com/youtube/v3/videos?'+(urllib.parse.urlencode(param))
            print('動画情報取得：合計' + str(item_count)+'件')
            print(target_url)
            req = urllib.request.Request(target_url)
            try:
                with urllib.request.urlopen(req) as res:
                    videos_body = json.load(res)
                    #CSV書き込み用データ準備
                    for item in videos_body['items']:
                        #値が存在しない場合ブランク
                        publishedAt = item['snippet']['publishedAt'] if 'publishedAt' in item['snippet'] else ''
                        title = item['snippet']['title'] if 'title' in item['snippet'] else ''
                        description = item['snippet']['description'] if 'description' in item['snippet'] else ''
                        url = item['id'] if 'id' in item else ''
                        thumbnail_url = item['snippet']['thumbnails']['high']['url'] if 'thumbnails' in item['snippet'] else ''
                        categoryId = item['snippet']['categoryId'] if 'categoryId' in item['snippet'] else ''
                        liveBroadcastContent = item['snippet']['liveBroadcastContent'] if 'liveBroadcastContent' in item['snippet'] else ''
                        if 'duration' in item['contentDetails']:
                            #durationを時分秒へ変換
                            duration = isodate.parse_duration(item['contentDetails']['duration'])
                        else:
                            duration = ''
                        viewCount = item['statistics']['viewCount'] if 'viewCount' in item['statistics'] else 0
                        likeCount = item['statistics']['likeCount'] if 'likeCount' in item['statistics'] else 0
                        favoriteCount = item['statistics']['favoriteCount'] if 'favoriteCount' in item['statistics'] else 0
                        commentCount = item['statistics']['commentCount'] if 'commentCount' in item['statistics'] else 0
                        embedHtml = item['player']['embedHtml'] if 'embedHtml' in item['player'] else ''
                        outputs.append([publishedAt, title, description, url, thumbnail_url, categoryId, liveBroadcastContent, duration, viewCount, likeCount, favoriteCount, commentCount, embedHtml])
                        n += 1
                    #CSV書き込み
                    for i in range(1, len(outputs)):
                        for j in range(1, 13):
                            ws.cell(row = i+1, column = j+1, value=outputs[i][j])
                    wb.save("data_"+channel_id+".xlsx") 
                    wb.close()
                    print(outputs)
                    
            except urllib.error.HTTPError as err:
                print(err)
                break
            except urllib.error.URLError as err:
                print(err)
                break
        
        #nextPageTokenが表示されなくなったらストップ
        if 'nextPageToken' in search_body:
            nextPageToken = search_body['nextPageToken']
        else:
            break
    except urllib.error.HTTPError as err:
        print(err)
        break
    except urllib.error.URLError as err:
        print(err)
        break